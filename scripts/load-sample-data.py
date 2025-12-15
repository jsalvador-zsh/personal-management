#!/usr/bin/env python3
"""
Script para cargar datos de prueba desde el Excel MODELO EXCEL ZABALAoriginaloriginal.xlsx
Este script extrae los datos y genera SQL para insertar en Supabase
"""

import openpyxl
from datetime import datetime
import json

# Cargar el archivo Excel
wb = openpyxl.load_workbook('../MODELO EXCEL ZABALAoriginaloriginal.xlsx')
ws = wb['BBD']

print("=== EXTRAYENDO DATOS DEL EXCEL ===\n")

# Extraer datos de la empresa
empresa_principal = "ELECTRODATA PERU"
cliente_principal = "COMPAÑIA MINERA ANTAMINA S.A."

print(f"Empresa: {empresa_principal}")
print(f"Cliente: {cliente_principal}\n")

# Extraer servicios (filas 98-112)
servicios = []
print("=== SERVICIOS ENCONTRADOS ===")
for i in range(98, 113):  # Filas 98-112
    row = list(ws[i])
    # Buscar el nombre del servicio en las celdas
    servicio_nombre = None
    for cell in row:
        if cell.value and isinstance(cell.value, str) and len(cell.value) > 10:
            if "Servicio" in cell.value or "Supervisión" in cell.value or "Administración" in cell.value:
                servicio_nombre = cell.value
                break

    if servicio_nombre:
        servicios.append(servicio_nombre)
        print(f"{len(servicios)}. {servicio_nombre}")

# Extraer trabajadores (filas 23-37)
trabajadores = []
print("\n=== TRABAJADORES ENCONTRADOS ===")
for i in range(23, 38):  # Filas 23-37
    row_values = [cell.value for cell in ws[i]]

    # Buscar fecha, empresa, nombre, servicio, puesto, habilitación
    fecha = None
    empresa = None
    nombre = None
    servicio = None
    puesto = None
    habilitacion = None

    for idx, val in enumerate(row_values):
        if isinstance(val, datetime):
            fecha = val
        elif val == cliente_principal:
            empresa = val
        elif val and isinstance(val, str):
            if "COMPAÑIA" not in val and "Servicio" not in val and len(val.split()) >= 2:
                if not nombre and all(c.isupper() or c.isspace() for c in val):
                    nombre = val
            if "Servicio" in str(val) or "Supervisión" in str(val) or "Administración" in str(val):
                servicio = val
            if val in ["ANALISTA DE INFRAESTRUCTURA", "TÉCNICO DE TELECOMUNICACIONES",
                       "SUPERVISOR DE SEGURIDAD", "SUPERVISOR DE OPERACIONES",
                       "GESTOR DE SERVICIOS", "ASISTENTE ADMINISTRATIVA", "ADMINISTRADOR DE CONTRATOS"]:
                puesto = val
            if val in ["Habilitado", "Inhabilitado"]:
                habilitacion = val

    if nombre and fecha:
        trabajadores.append({
            "nombre": nombre,
            "fecha_evaluacion": fecha.strftime("%Y-%m-%d") if fecha else None,
            "empresa": empresa,
            "servicio": servicio,
            "puesto": puesto,
            "habilitacion": habilitacion
        })
        print(f"{len(trabajadores)}. {nombre} - {puesto} - {habilitacion}")

print(f"\nTotal trabajadores: {len(trabajadores)}")
print(f"Total servicios: {len(servicios)}")

# Generar SQL
print("\n=== GENERANDO SQL ===")

sql_output = """-- ============================================================================
-- DATOS DE PRUEBA BASADOS EN MODELO EXCEL ZABALA
-- ============================================================================

-- ============================================================================
-- EMPRESA (Cliente Principal)
-- ============================================================================
INSERT INTO companies (name, description, work_mode, cost_center) VALUES
('COMPAÑIA MINERA ANTAMINA S.A.', 'Unidad minera principal - Cliente', 'presencial', 'OP 2212-035');

-- ============================================================================
-- SERVICIOS
-- ============================================================================
"""

for idx, servicio in enumerate(servicios, 1):
    # Limpiar nombre del servicio (máximo 255 caracteres)
    servicio_clean = servicio.strip()[:255].replace("'", "''")
    sql_output += f"""INSERT INTO services (name, company_id, status, description) VALUES
('{servicio_clean}',
 (SELECT id FROM companies WHERE name = 'COMPAÑIA MINERA ANTAMINA S.A.' LIMIT 1),
 'activo',
 'Servicio de telecomunicaciones y soporte técnico');
"""

sql_output += """
-- ============================================================================
-- CURSOS (Ejemplos basados en el contexto minero)
-- ============================================================================
INSERT INTO courses (name, description, duration_hours) VALUES
('Seguridad en Operaciones Mineras', 'Curso obligatorio de seguridad para operaciones en mina', 40),
('Telecomunicaciones en Minería', 'Fundamentos de sistemas de telecomunicaciones en entornos mineros', 32),
('Mantenimiento de Equipos de Flota Pesada', 'Procedimientos de mantenimiento preventivo y correctivo', 48),
('Gestión de Servidores y Bases de Datos', 'Administración de infraestructura IT en entornos críticos', 56),
('Ciberseguridad Industrial', 'Protección de sistemas de control industrial', 24),
('Itrack y Sistemas de Monitoreo', 'Operación y mantenimiento de sistemas de tracking', 16);

-- ============================================================================
-- TRABAJADORES
-- ============================================================================
"""

for trabajador in trabajadores:
    # Separar nombre en partes
    partes_nombre = trabajador["nombre"].split()
    if len(partes_nombre) >= 2:
        nombres = partes_nombre[-2] + " " + partes_nombre[-1]  # Los últimos dos son nombres
        apellidos = " ".join(partes_nombre[:-2])  # El resto son apellidos
    else:
        nombres = trabajador["nombre"]
        apellidos = ""

    # Generar DNI ficticio (8 dígitos)
    dni = f"{70000000 + trabajadores.index(trabajador):08d}"

    # Generar email
    email_nombre = nombres.lower().replace(" ", ".")
    email_apellido = apellidos.lower().split()[0] if apellidos else ""
    email = f"{email_nombre}.{email_apellido}@electrodata.com.pe" if email_apellido else f"{email_nombre}@electrodata.com.pe"

    # Estado
    status = "habilitado" if trabajador["habilitacion"] == "Habilitado" else "inhabilitado"

    # Puesto (usar como position)
    puesto = trabajador["puesto"] if trabajador["puesto"] else "TÉCNICO GENERAL"

    sql_output += f"""INSERT INTO workers (full_name, dni, email, phone, position, status, company_id) VALUES
('{trabajador["nombre"]}', '{dni}', '{email}', '+51 900 000 {str(trabajadores.index(trabajador)+100)[:3]}',
 '{puesto}', '{status}',
 (SELECT id FROM companies WHERE name = 'COMPAÑIA MINERA ANTAMINA S.A.' LIMIT 1));
"""

sql_output += """
-- ============================================================================
-- CERTIFICACIONES (Ejemplos)
-- ============================================================================
-- Asignar certificaciones a los primeros 10 trabajadores
DO $$
DECLARE
    worker_record RECORD;
    course_record RECORD;
    issue_date DATE;
    expiry_date DATE;
BEGIN
    FOR worker_record IN (SELECT id FROM workers LIMIT 10)
    LOOP
        FOR course_record IN (SELECT id FROM courses ORDER BY RANDOM() LIMIT 2)
        LOOP
            issue_date := CURRENT_DATE - (RANDOM() * 365)::INTEGER;
            expiry_date := issue_date + INTERVAL '1 year';

            INSERT INTO certifications (worker_id, course_id, issue_date, expiry_date)
            VALUES (worker_record.id, course_record.id, issue_date, expiry_date);
        END LOOP;
    END LOOP;
END $$;

-- ============================================================================
-- EVALUACIONES (Ejemplos)
-- ============================================================================
DO $$
DECLARE
    worker_record RECORD;
    evaluator_id UUID;
    eval_types TEXT[] := ARRAY['admin', 'medico', 'rrhh'];
    eval_type TEXT;
BEGIN
    -- Obtener un evaluador (el primer usuario admin)
    SELECT id INTO evaluator_id FROM users WHERE role = 'admin' LIMIT 1;

    IF evaluator_id IS NOT NULL THEN
        FOR worker_record IN (SELECT id FROM workers LIMIT 10)
        LOOP
            eval_type := eval_types[1 + FLOOR(RANDOM() * 3)::INTEGER];

            INSERT INTO evaluations (worker_id, evaluator_id, evaluation_type, score, date, comments)
            VALUES (
                worker_record.id,
                evaluator_id,
                eval_type::evaluation_type,
                70 + (RANDOM() * 30)::INTEGER,
                CURRENT_DATE - (RANDOM() * 180)::INTEGER,
                'Evaluación realizada correctamente. Personal cumple con estándares.'
            );
        END LOOP;
    END IF;
END $$;

-- ============================================================================
-- HOMOLOGACIONES (Cursos obligatorios por empresa)
-- ============================================================================
DO $$
DECLARE
    company_id UUID;
    course_record RECORD;
BEGIN
    SELECT id INTO company_id FROM companies WHERE name = 'COMPAÑIA MINERA ANTAMINA S.A.' LIMIT 1;

    IF company_id IS NOT NULL THEN
        FOR course_record IN (SELECT id FROM courses WHERE name LIKE '%Seguridad%' OR name LIKE '%Ciberseguridad%')
        LOOP
            INSERT INTO homologations (company_id, course_id, is_required)
            VALUES (company_id, course_record.id, true)
            ON CONFLICT DO NOTHING;
        END LOOP;
    END IF;
END $$;

COMMIT;
"""

# Guardar SQL
with open('../supabase/migrations/sample-data.sql', 'w', encoding='utf-8') as f:
    f.write(sql_output)

print("✓ Archivo SQL generado: supabase/migrations/sample-data.sql")
print("\nPara cargar los datos:")
print("1. Abre el SQL Editor en Supabase")
print("2. Copia el contenido de 'supabase/migrations/sample-data.sql'")
print("3. Ejecuta el script")
print("\n¡Listo! Los datos de prueba estarán cargados en tu base de datos.")
